import tkinter as tk
import subprocess
import openpyxl
from openpyxl import Workbook  # Correct import
import os
from openpyxl.styles import Font


# Global variables
NPV_entries = []  # Changed CFV_entries to NPV_entries
payment_year_spinbox = None
payment_patterns_spinboxes = []
discount_rate_spinboxes = []
year_spinboxs = []
amount_entrys = []
amount_risk_entrys = []
amount_out_entrys = []
year_amount_spinbox = []
year_risk_spinboxs = []
year_out_spinboxs = []


def calculate_discounting_factor(period: int, discount_rate: float) -> float:
    return 1 / (1 + discount_rate) ** period


def calculate_discount_factors():
    try:
        payment_year = int(payment_year_spinbox.get())
        discount_rates = [float(spinbox.get()) for spinbox in discount_rate_spinboxes]

        # Get the input values
        estimated_claim_amount = float(Estimated_C_amount_entry.get())
        Outstanding_Claims = float(Outstanding_Claims_entry.get())
        Risk_Adjustment = float(Risk_Adjustment_entry.get())

        # Get payment patterns
        payment_patterns = [float(spinbox.get()) for spinbox in payment_patterns_spinboxes]

        total_npv = 0.0

        # Calculate NPV for each year
        for i in range(payment_year):
            if i < len(NPV_entries):
                NPV_entries[i].config(state="normal")
                NPV_entries[i].delete(0, tk.END)

                # Calculate individual payments for each component
                estimated_payment = estimated_claim_amount * payment_patterns[i]
                outstanding_claims_payment = Outstanding_Claims * payment_patterns[i]
                risk_adjustment_payment = Risk_Adjustment * payment_patterns[i]

                # Calculate total payment for this period
                total_payment = estimated_payment + outstanding_claims_payment + risk_adjustment_payment

                # Calculate discount factor and NPV
                discount_factor = calculate_discounting_factor(i + 1, discount_rates[i])
                npv = total_payment * discount_factor

                # Update NPV entry
                NPV_entries[i].insert(0, f"{npv:.2f}")
                NPV_entries[i].config(state="readonly")

                total_npv += npv

        # Update total NPV entry box
        Total_NPV_entry.config(state="normal")
        Total_NPV_entry.delete(0, tk.END)
        Total_NPV_entry.insert(0, f"{total_npv:.2f}")
        Total_NPV_entry.config(state="readonly")

    except ValueError as e:
        print(f"Error: Please enter valid numbers. {str(e)}")


def update_npv():
    """Updates the NPV when the payment pattern or discount rate spinboxes are changed."""
    calculate_discount_factors()  # Recalculate the NPV whenever an input changes.


def generate_discount_rate_spinboxes():
    """Generate spinboxes for discount rates based on the selected payment year."""
    for spinbox in discount_rate_spinboxes:
        spinbox.destroy()  # Remove previous discount rate spinboxes

    discount_rate_spinboxes.clear()

    payment_year = int(payment_year_spinbox.get())

    for x in range(payment_year):
        spinbox = tk.Spinbox(
            user_info_frame, from_=0.0, to=1.0, increment=0.001, width=10, format="%.7f", command=update_npv
        )
        spinbox.grid(row=24, column=1 + x)
        discount_rate_spinboxes.append(spinbox)


def generate_payment_pattern_spinboxes():
    """Generate spinboxes for payment patterns based on the selected payment year and dynamically create NPV entries."""
    # Clear previous spinboxes and entries
    for spinbox in payment_patterns_spinboxes:
        spinbox.destroy()
    payment_patterns_spinboxes.clear()

    for entry in NPV_entries:
        entry.destroy()
    NPV_entries.clear()

    payment_year = int(payment_year_spinbox.get())

    for x in range(payment_year):
        spinbox = tk.Spinbox(
            user_info_frame, from_=0.0, to=1.0, increment=0.0001, width=10, format="%.4f", command=update_npv
        )
        spinbox.grid(row=26, column=1 + x)
        payment_patterns_spinboxes.append(spinbox)

        # Create corresponding NPV entry box
        npv_entry = tk.Entry(user_info_frame, width=20)
        npv_entry.grid(row=27, column=1 + x)
        NPV_entries.append(npv_entry)
        npv_entry.config(state="readonly")

    # Also generate discount rate spinboxes when payment years are updated
    generate_discount_rate_spinboxes()


def export_to_excel(total_npv):
    # Specify the Excel file path
    excel_file_path = r"C:\GI_Automation\\output\user_data.xlsx"

    # Try to load the workbook, or create a new one if it doesn't exist
    if os.path.exists(excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
    else:
        print(f"The file {excel_file_path} does not exist.")
        return

    if "User Data" not in wb.sheetnames:
        user_data_sheet = wb.create_sheet("User Data")
    else:
        user_data_sheet = wb["User Data"]

    Total_NPV_value = float(Total_NPV_entry.get())

    next_column = 2  # Change this as per your requirement

    # Write total NPV value to row 5
    user_data_sheet.cell(row=5, column=next_column, value=Total_NPV_value)

    # Get the product name from the entry widget
    product_name = product_name_entry.get()

    # Check if the sheet with the product name exists, otherwise create it with "old" appended
    if product_name not in wb.sheetnames:
        # Create a new sheet with the name from product_name_entry
        product_sheet = wb.create_sheet(f"{product_name} new")
        product_sheet.tab_color = "FF0000"  # Set tab color to red
    else:
        # If the sheet already exists, create a new sheet with the name from product_name_entry + "old"
        new_sheet_name = f"{product_name} new"
        product_sheet = wb.create_sheet(new_sheet_name)
        product_sheet.tab_color = "FF0000"  # Set tab color to red

    # [Rest of the function remains the same...]    # Get the product name from the entry widget
    product_name = product_name_entry.get()

    

    # Labels for calculations in the sheet
    product_sheet.cell(row=3, column=1, value="Estimated Claim Amount").font = Font(bold=True)

    product_sheet.cell(row=4, column=1, value="Year").font = Font(bold=True)
    product_sheet.cell(row=4, column=2, value="Amount").font = Font(bold=True)
    product_sheet.cell(row=10, column=1, value="Estimated Claim Amount Total")
    product_sheet.cell(row=3, column=4, value="Risk Adjustment").font = Font(bold=True)
    product_sheet.cell(row=10, column=4, value="Risk Adjustment Total")
    product_sheet.cell(row=3, column=7, value="Outstanding Claims").font = Font(bold=True)
    product_sheet.cell(row=10, column=7, value="Outstanding Claims Total")
    product_sheet.cell(row=13, column=1, value="Discount Rates")
    product_sheet.cell(row=14, column=1, value="Payment Year")
    product_sheet.cell(row=17, column=1, value="Payment Patterns")
    product_sheet.cell(row=19, column=1, value="Estimated payment")
    product_sheet.cell(row=20, column=1, value="Risk Adjustment payment")
    product_sheet.cell(row=21, column=1, value="Outstanding Claims payment")
    product_sheet.cell(row=22, column=1, value="Discount factor")
    product_sheet.cell(row=27, column=1, value="Total NPV new")
    product_sheet.cell(row=28, column=1, value="NPV new")

    product_sheet.cell(row=4, column=4, value="=A4").font = Font(bold=True)
    product_sheet.cell(row=4, column=5, value="=B4").font = Font(bold=True)
    product_sheet.cell(row=4, column=7, value="=A4").font = Font(bold=True)
    product_sheet.cell(row=4, column=8, value="=B4").font = Font(bold=True)

    # Gather data from entry widgets
    estimated_claim_amount = Estimated_C_amount_entry.get()
    Risk_Adjustment = Risk_Adjustment_entry.get()
    Outstanding_Claims = Outstanding_Claims_entry.get()
    discount_rates = [spinbox.get() for spinbox in discount_rate_spinboxes]
    payment_years = int(payment_year_spinbox.get())
    payment_patterns = [spinbox.get() for spinbox in payment_patterns_spinboxes]

    # Insert the values into the sheet
    product_sheet.cell(row=10, column=2, value=estimated_claim_amount)
    product_sheet.cell(row=10, column=5, value=Risk_Adjustment)
    product_sheet.cell(row=10, column=8, value=Outstanding_Claims)

    year_amount = [spinbox.get() for spinbox in year_amount_spinbox]
    year_risk_amount = [spinbox.get() for spinbox in year_amount_spinbox]
    year_out_amount = [spinbox.get() for spinbox in year_amount_spinbox]

    amount_value = [entry.get() for entry in amount_entrys]
    risk_amount_value = [entry.get() for entry in amount_risk_entrys]
    out_amount_value = [entry.get() for entry in amount_out_entrys]

    # Write amount, risk, and out amounts to the "Product Data" sheet
    for i, value in enumerate(amount_value):
        product_sheet.cell(row=5 + i, column=2, value=value)  # Amount in column 2
    for i, value in enumerate(risk_amount_value):
        product_sheet.cell(row=5 + i, column=5, value=value)  # Risk Amount in column 5
    for i, value in enumerate(out_amount_value):
        product_sheet.cell(row=5 + i, column=8, value=value)  # Out Amount in column 8

    for i, value in enumerate(year_amount):
        product_sheet.cell(row=5 + i, column=1, value=value)
    for i, value in enumerate(year_risk_amount):
        product_sheet.cell(row=5 + i, column=4, value=value)
    for i, value in enumerate(year_out_amount):
        product_sheet.cell(row=5 + i, column=7, value=value)

    # Insert discount rates and payment years
    for i, value in enumerate(discount_rates):
        product_sheet.cell(row=13, column=2 + i, value=value)  # Discount rates
    for i in range(payment_years):
        product_sheet.cell(row=14, column=2 + i, value=i + 1)  # Payment years
    for i, value in enumerate(payment_patterns):
        product_sheet.cell(row=17, column=2 + i, value=value)  # Payment patterns

    # Insert NPV values (from entry widgets)
    npv_values = [entry.get() for entry in NPV_entries]  # Get all NPV entry values
    for i, value in enumerate(npv_values):
        product_sheet.cell(row=27, column=2 + i, value=value)

    # Write Total NPV to the sheet
    product_sheet.cell(row=28, column=2, value=total_npv)
    # Set up formulas for columns B-O
    for i in range(2, 16):
        column_letter = chr(64 + i)
        product_sheet.cell(row=19, column=i, value=f'=IF(B10*{column_letter}17=0, "", B10*{column_letter}17)')
        product_sheet.cell(row=20, column=i, value=f'=IF(E10*{column_letter}17=0, "", E10*{column_letter}17)')
        product_sheet.cell(row=21, column=i, value=f'=IF(H10*{column_letter}17=0, "", H10*{column_letter}17)')
        product_sheet.cell(
            row=22,
            column=i,
            value=f'=IF(SUM({column_letter}19:{column_letter}21)=0, "", {column_letter}24/SUM({column_letter}19:{column_letter}21))',
        )

    for i in range(14):
        next_column_letter = chr(66 + i)  # Start from 'B' (ASCII value for 'B' is 66)

        # Insert the corrected formula into the cell
        product_sheet.cell(
            row=22,
            column=i + 2,  # Ensures starting from column B
            value=f'=IF({next_column_letter}14="", "", (1/(1+${next_column_letter}$13)^{next_column_letter}14))',
        )
    # Set formulas for calculations in rows 9 to 12

    # Save the Excel workbook
    wb.save(excel_file_path)
    print("Data exported successfully to Excel.")


def calculate_totals():
    # Calculate total for amount_entrys
    total_amount = 0
    for entry in amount_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_amount += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Calculate total for amount_out_entrys
    total_outstanding = 0
    for entry in amount_out_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_outstanding += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Calculate total for amount_risk_entrys
    total_risk = 0
    for entry in amount_risk_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_risk += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Update each entry with the calculated totals
    Estimated_C_amount_entry.delete(0, tk.END)
    Estimated_C_amount_entry.insert(0, str(total_amount))

    Risk_Adjustment_entry.delete(0, tk.END)
    Risk_Adjustment_entry.insert(0, str(total_risk))

    Outstanding_Claims_entry.delete(0, tk.END)
    Outstanding_Claims_entry.insert(0, str(total_outstanding))


def generate_year():
    for spinbox in year_amount_spinbox:
        spinbox.destroy()
    year_amount_spinbox.clear()

    for spinbox in year_risk_spinboxs:
        spinbox.destroy()
    year_risk_spinboxs.clear()

    for spinbox in year_out_spinboxs:
        spinbox.destroy()
    year_out_spinboxs.clear()

    for entry in amount_entrys:
        entry.destroy()
    amount_entrys.clear()

    for entry in amount_risk_entrys:
        entry.destroy()
    amount_risk_entrys.clear()

    for entry in amount_out_entrys:
        entry.destroy()
    amount_out_entrys.clear()

    # Convert year to an integer
    year = int(year_spinbox.get())

    for x in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=2 + x, column=0)
        year_amount_spinbox.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=2 + x, column=1)
        amount_entrys.append(entry)

    for y in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=8 + y, column=0)
        year_risk_spinboxs.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=8 + y, column=1)
        amount_risk_entrys.append(entry)

    for a in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=14 + a, column=0)
        year_out_spinboxs.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=14 + a, column=1)
        amount_out_entrys.append(entry)


def next_interface():
    try:
        # Get Total NPV value
        Total_NPV = Total_NPV_entry.get()

        excel_file_path = r"C:\GI_Automation\\output\user_data.xlsx"

        if not os.path.exists(excel_file_path):

            return

        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)

        # Get the User Data sheet
        if "User Data" in wb.sheetnames:
            user_data_sheet = wb["User Data"]

            # Add SUM formulas to column Q (17)
            user_data_sheet.cell(row=5, column=17, value="=SUM(A5:P5)")

            if isinstance(user_data_sheet, (int, float)):
                print()
            else:
                print("The value is not a number or has not been calculated.")

            # Save and close the workbook
            wb.save(excel_file_path)
            wb.close()

        # Export the NPV value
        export_to_excel(Total_NPV)

        # Close current interface and open new one
        interface.destroy()
        subprocess.Popen(["python", r"C:\GI_Automation\GUI\final.py"])

    except Exception as e:
        print(f"Error in next_interface: {e}")

def back():

    interface.destroy()
    subprocess.Popen(["python", r"C:\GI_Automation\GUI\interface3_1.py"])


# Setup main window
interface = tk.Tk()
interface.title("Interface2")
interface.geometry("2500x1000")


frame = tk.Frame(interface)
frame.pack()

user_info_frame = tk.LabelFrame(frame, text="Subsq new rate ", padx=10, pady=10)
user_info_frame.grid(row=0, column=0, padx=10, pady=10)


product_name_lable = tk.Label(user_info_frame, text="product :")
product_name_lable.grid(row=1, column=0, sticky="w")


product_name_entry = tk.Entry(user_info_frame, width=20)
product_name_entry.grid(row=1, column=1, pady=5)

year_lable = tk.Label(user_info_frame, text="Year:")
year_lable.grid(row=1, column=3, sticky="w")

year_spinbox = tk.Spinbox(user_info_frame, from_=1, to=5, command=generate_year)
year_spinbox.grid(row=1, column=4, pady=5)

# Estimated Claim Amount
Estimated_C_amount_label = tk.Label(user_info_frame, text="Estimated Claim Amount (IBNR):")
Estimated_C_amount_label.grid(row=7, column=0, sticky="w")

Risk_Adjustment_label = tk.Label(user_info_frame, text="Risk Adjustment:")
Risk_Adjustment_label.grid(row=13, column=0, sticky="w")

Outstanding_Claims_label = tk.Label(user_info_frame, text="Outstanding Claims:")
Outstanding_Claims_label.grid(row=20, column=0, sticky="w")


Estimated_C_amount_entry = tk.Entry(user_info_frame, width=20)
Estimated_C_amount_entry.grid(row=7, column=1, pady=5)
Risk_Adjustment_entry = tk.Entry(user_info_frame, width=20)
Risk_Adjustment_entry.grid(row=13, column=1, pady=5)
Outstanding_Claims_entry = tk.Entry(user_info_frame, width=20)
Outstanding_Claims_entry.grid(row=20, column=1, pady=5)


generate_button = tk.Button(user_info_frame, text="Generate calculation ", command=calculate_totals)
generate_button.grid(row=20, column=2, pady=10)


# Payment Year (single spinbox)
Payment_year_label = tk.Label(user_info_frame, text="Payment Year:")
Payment_year_label.grid(row=23, column=0, sticky="w")

payment_year_spinbox = tk.Spinbox(user_info_frame, from_=1, to=12, command=generate_payment_pattern_spinboxes)
payment_year_spinbox.grid(row=23, column=1, pady=5)

# Discount Rate (dynamic spinboxes will be created)
Discount_rate_label = tk.Label(user_info_frame, text="Discount Rate (%):")
Discount_rate_label.grid(row=24, column=0, sticky="w")

# Estimated Payment Pattern
E_pay_label = tk.Label(user_info_frame, text="Estimated Payment Pattern (%):")
E_pay_label.grid(row=26, column=0, sticky="w")

# Total Net Present Value new
Total_NPV_label = tk.Label(user_info_frame, text="Total Net Present Value new:")
Total_NPV_label.grid(row=28, column=0, sticky="w")

Total_NPV_entry = tk.Entry(user_info_frame, width=20)
Total_NPV_entry.grid(row=28, column=1)
Total_NPV_entry.config(state="readonly")  # Set to readonly initially

# Net Present Value new
NPV_label = tk.Label(user_info_frame, text="Net Present Value new:")
NPV_label.grid(row=27, column=0, sticky="w")
calculate_button = tk.Button(user_info_frame, text="Calculate", command=calculate_discount_factors)
calculate_button.grid(row=35, column=0, pady=10)

export_button = tk.Button(
    user_info_frame,
    text="Export to Excel",
    command=lambda: export_to_excel(Total_NPV_entry.get()),
)
export_button.grid(row=35, column=3, pady=20)
# Next and Back Buttons
next_button = tk.Button(user_info_frame, text="Next", command=next_interface)
next_button.grid(row=35, column=1, pady=10)

# Back Button
back_button = tk.Button(user_info_frame, text="Back", command=back)
back_button.grid(row=35, column=2, pady=10)


# Start main loop
interface.mainloop()
