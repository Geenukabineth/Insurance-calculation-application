import os
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook


def generated_excel():
    try:
        # Define the file path
        file_path = r"C:\GI_Automation\\output\user_data.xlsx"

        # Load the existing workbook or create a new one if it doesn't exist
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws_calculations = wb.active  # Get the active worksheet
        else:
            wb = Workbook()
            ws_calculations = wb.active
            ws_calculations.title = "Calculations"

        # Hide specific rows
        for row in [2, 3, 4, 5]:
            ws_calculations.row_dimensions[row].hidden = True

        # Define the calculations with proper numeric values
        calculations = [
            ("Opening claim liability", "=Q2"),
            ("Interest accretion", "= Q3"),
            ("Changes in estimation", "= Q2 - Q4"),
            ("Changes in financial assumptions", "= Q5 - Q4"),
            ("As Claim Liability per Movement", "=B7+B8+B9+B10"),
            ("Closing Claim liability as estimation", "=Q5"),
            ("Diff", "=B10-B11"),
        ]

        # Write calculations to worksheet
        for i, (label, value) in enumerate(calculations, start=7):
            ws_calculations.cell(row=i, column=1, value=label)
            if isinstance(value, str) and value.startswith("="):
                ws_calculations.cell(row=i, column=2, value=value)
            else:
                ws_calculations.cell(row=i, column=2, value=float(value))

        # Save the workbook
        try:
            wb.save(file_path)
            print(f"Excel file successfully updated at {file_path}")
            os.startfile(file_path)
        except PermissionError:
            messagebox.showerror("Error", "Cannot save the file. Please close Excel and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving file: {str(e)}")

    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")


def create_gui():
    # Create the main Tkinter window
    root = tk.Tk()
    root.title("Final Output")
    root.geometry("300x150")

    # Create a frame for better organization
    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(expand=True)

    # Button to generate Excel data
    fetch_button = tk.Button(frame, text="Generate Excel Data", command=generated_excel, width=20, height=2)
    fetch_button.pack(pady=10)

    return root


if __name__ == "__main__":
    root = create_gui()
    root.mainloop()
