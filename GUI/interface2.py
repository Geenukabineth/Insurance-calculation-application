import tkinter as tk
import subprocess
import openpyxl
from openpyxl import Workbook
import os
from openpyxl.styles import Font


Estimated_C_amount_entrys = []
discount_rate_spinboxes = []
payment_patterns_spinboxes = []
Risk_Adjustment_entrys = []
Outstanding_Claims_entrys = []
CFV_entries = []
Interest_accretion_entries = []
year_spinboxs = []
amount_entrys = []
amount_risk_entrys = []
amount_out_entrys = []
year_amount_spinbox = []
year_risk_spinboxs = []
year_out_spinboxs = []


def calculate_discounting_factor(period: int, discount_rate: float) -> float:
    return 1 / (1 + discount_rate) ** period


def calculate_discount_factors():
    try:
        payment_year = int(payment_year_spinbox.get())
        discount_rates = [float(spinbox.get()) for spinbox in discount_rate_spinboxes]
        
        estimated_claim_amount = float(Estimated_C_amount_entry.get())
        Risk_Adjustment_amount = float(Risk_Adjustment_entry.get())
        Outstanding_Claims_amount = float(Outstanding_Claims_entry.get())
        
        payment_patterns = [float(spinbox.get()) for spinbox in payment_patterns_spinboxes]
        
        total_cfv = 0
        total_interest_accretion = 0

        # Calculate CFV for each year
        for i in range(payment_year):
            if i < len(CFV_entries):
                CFV_entries[i].config(state="normal")
                CFV_entries[i].delete(0, tk.END)
                
                # Calculate estimated payments for each component
                estimated_payment = estimated_claim_amount * payment_patterns[i]
                risk_adjustment_payment = Risk_Adjustment_amount * payment_patterns[i]
                outstanding_claims_payment = Outstanding_Claims_amount * payment_patterns[i]
                
                # Total payment for this period
                total_payment = estimated_payment + risk_adjustment_payment + outstanding_claims_payment
                
                # Calculate discount factor
                discount_factor = calculate_discounting_factor(i + 1, discount_rates[i])
                
                # Calculate CFV
                cfv_value = total_payment * discount_factor
                
                CFV_entries[i].insert(0, f"{cfv_value:.2f}")
                CFV_entries[i].config(state="readonly")
                total_cfv += cfv_value

        # Update total CFV field
        total_cfv_entry.config(state="normal")
        total_cfv_entry.delete(0, tk.END)
        total_cfv_entry.insert(0, f"{total_cfv:.2f}")
        total_cfv_entry.config(state="readonly")

        # Calculate Interest Accretion
        for i in range(payment_year):
            if i < len(Interest_accretion_entries):
                Interest_accretion_entries[i].config(state="normal")
                Interest_accretion_entries[i].delete(0, tk.END)
                
                # Get CFV value for this period
                cfv = float(CFV_entries[i].get())
                
                # Calculate Interest Accretion
                interest_accretion = cfv * (1 + discount_rates[i])
                
                Interest_accretion_entries[i].insert(0, f"{interest_accretion:.2f}")
                Interest_accretion_entries[i].config(state="readonly")
                total_interest_accretion += interest_accretion

        # Update total Interest Accretion field
        total_interest_accretion_entry.config(state="normal")
        total_interest_accretion_entry.delete(0, tk.END)
        total_interest_accretion_entry.insert(0, f"{total_interest_accretion:.2f}")
        total_interest_accretion_entry.config(state="readonly")

    except ValueError as e:
        print(f"Error: Please enter valid numbers. {str(e)}")
def update_cfv_and_interest_accretion():
    """Updates the CFV and Interest Accretion when the payment pattern or discount rate spinboxes are changed."""
    calculate_discount_factors()  # Recalculate the CFV and Interest Accretion whenever an input changes.


def generate_discount_rate_spinboxes():
    """Generate spinboxes for discount rates based on the selected payment year."""
    for spinbox in discount_rate_spinboxes:
        spinbox.destroy()  # Remove previous discount rate spinboxes

    discount_rate_spinboxes.clear()

    payment_year = int(payment_year_spinbox.get())

    for x in range(payment_year):
        spinbox = tk.Spinbox(
            user_info_frame,
            from_=0.0,
            to=1.0,
            increment=0.001,
            width=10,
            format="%.7f",
            command=update_cfv_and_interest_accretion,
        )
        spinbox.grid(row=24, column=1 + x)
        discount_rate_spinboxes.append(spinbox)


def export_to_excel(total_cfv, total_interest_accretion):
    # Specify the Excel file path
    excel_file_path = r"C:\GI_Automation\output\\user_data.xlsx"

    # Try to load the workbook, or create a new one if it doesn't exist
    if not os.path.exists(excel_file_path):
        wb = Workbook()  # Create a new workbook
        ws = wb.active
        ws.title = "User Data"
    else:
        wb = openpyxl.load_workbook(excel_file_path)  # Load existing workbook
        ws = wb.active  
        # Get the active sheet

    if "User Data" not in wb.sheetnames:
        user_data_sheet = wb.create_sheet("User Data")
    else:
        user_data_sheet = wb["User Data"]

    # Get the total CFV and total interest accretion values from the Entry widgets
    total_cfv_value = float(total_cfv_entry.get())
    total_interest_accretion_value = float(total_interest_accretion_entry.get())

    
    next_column = 2
    user_data_sheet.cell(row=2, column=next_column, value=total_cfv_value)
    # Move to the next column for total interest accretion
    user_data_sheet.cell(row=3, column=next_column, value=total_interest_accretion_value)

    # Get the product name from the entry widget
    product_name = product_name_entry.get()

    # Hide the first two rows in the "User Data" sheet (if you want to hide rows in this sheet)

    # Check if the sheet with the product name exists, otherwise create it
    if product_name not in wb.sheetnames:
        # Create a new sheet with the name from product_name_entry
        initial_ws = wb.create_sheet(product_name)
        initial_ws.tab_color = "FF0000"  # Set tab color to red
    else:
        # If the sheet already exists, access it
        initial_ws = wb[product_name]

    # Set up headers in the initial_ws sheet
    initial_ws.cell(row=3, column=1, value="Estimated Claim Amount").font = Font(bold=True)
    initial_ws.cell(row=2, column=1, value="Product").font = Font(bold=True)
    initial_ws.cell(row=4, column=1, value="Year").font = Font(bold=True)
    initial_ws.cell(row=4, column=2, value="Amount").font = Font(bold=True)
    initial_ws.cell(row=10, column=1, value="Estimated Claim Amount Total")
    initial_ws.cell(row=3, column=4, value="Risk Adjustment ").font = Font(bold=True)
    initial_ws.cell(row=10, column=4, value="Risk Adjustment Total")
    initial_ws.cell(row=3, column=7, value="Outstanding Claims ").font = Font(bold=True)
    initial_ws.cell(row=10, column=7, value="Outstanding Claims Total")
    initial_ws.cell(row=13, column=1, value="Discount Rates")
    initial_ws.cell(row=14, column=1, value="Payment Year")
    initial_ws.cell(row=17, column=1, value="Payment Patterns")
    initial_ws.cell(row=19, column=1, value="Estimated payment")
    initial_ws.cell(row=20, column=1, value="Risk Adjustment payment")
    initial_ws.cell(row=21, column=1, value="Outstanding Claims payment")
    initial_ws.cell(row=22, column=1, value="Discount factor")
    initial_ws.cell(row=24, column=1, value="CFV")
    initial_ws.cell(row=25, column=1, value="Interest Accretion")
    initial_ws.cell(row=27, column=1, value="Total CFV")
    initial_ws.cell(row=28, column=1, value="Total Interest Accretion")
    initial_ws.cell(row=29, column=1, value="Difference")

    # Initial hidden calculations
    initial_ws.cell(row=29, column=2, value="=B27-B28")

    initial_ws.cell(row=4, column=4, value="=A4").font = Font(bold=True)

    initial_ws.cell(row=4, column=5, value="=B4").font = Font(bold=True)

    initial_ws.cell(row=4, column=7, value="=A4").font = Font(bold=True)

    initial_ws.cell(row=4, column=8, value="=B4").font = Font(bold=True)

    # Set up formulas for columns B-O
    for i in range(2, 16):
        column_letter = chr(64 + i)
        initial_ws.cell(row=19, column=i, value=f'=IF(B10*{column_letter}17=0, "", B10*{column_letter}17)')
        initial_ws.cell(row=20, column=i, value=f'=IF(E10*{column_letter}17=0, "", E10*{column_letter}17)')
        initial_ws.cell(row=21, column=i, value=f'=IF(H10*{column_letter}17=0, "", H10*{column_letter}17)')
        initial_ws.cell(
            row=22,
            column=i,
            value=f'=IF(SUM({column_letter}19:{column_letter}21)=0, "", {column_letter}24/SUM({column_letter}19:{column_letter}21))',
        )

        product_name = product_name_entry.get()
        initial_ws.cell(row=2, column=2, value=product_name)

        # Next column calculation (avoid out-of-bounds)

    for i in range(14):
        next_column_letter = chr(66 + i)  # Start from 'B' (ASCII value for 'B' is 66)

        # Insert the corrected formula into the cell
        initial_ws.cell(
            row=22,
            column=i + 2,  # Ensures starting from column B
            value=f'=IF({next_column_letter}14="", "", (1/(1+${next_column_letter}$13)^{next_column_letter}14))',
        )
    # Gather data for the "Initial Estimation" sheet
    estimated_claim_amount = Estimated_C_amount_entry.get()
    Risk_Adjustment_amount = Risk_Adjustment_entry.get()
    Outstanding_Claims_amount = Outstanding_Claims_entry.get()
    payment_year = int(payment_year_spinbox.get())
    discount_rates = [spinbox.get() for spinbox in discount_rate_spinboxes]
    payment_patterns = [spinbox.get() for spinbox in payment_patterns_spinboxes]

    year_amount = [spinbox.get() for spinbox in year_amount_spinbox]
    year_risk_amount = [spinbox.get() for spinbox in year_amount_spinbox]
    year_out_amount = [spinbox.get() for spinbox in year_amount_spinbox]

    amount_value = [entry.get() for entry in amount_entrys]
    risk_amount_value = [entry.get() for entry in amount_risk_entrys]
    out_amount_value = [entry.get() for entry in amount_out_entrys]

    # Write amount values to Excel (each value in its own row)
    for i, value in enumerate(amount_value):
        initial_ws.cell(row=5 + i, column=2, value=value)  # Amount in column 2

    # Write risk amount values to Excel
    for i, value in enumerate(risk_amount_value):
        initial_ws.cell(row=5 + i, column=5, value=value)  # Risk Amount in column 5

    # Write out amount values to Excel
    for i, value in enumerate(out_amount_value):
        initial_ws.cell(row=5 + i, column=8, value=value)  # Out Amount in column 8

    for i, value in enumerate(year_amount):
        initial_ws.cell(row=5 + i, column=1, value=value)
    for i, value in enumerate(year_risk_amount):
        initial_ws.cell(row=5 + i, column=4, value=value)
    for i, value in enumerate(year_out_amount):
        initial_ws.cell(row=5 + i, column=7, value=value)

        # Write data to "Initial Estimation" sheet
    initial_ws.cell(row=10, column=2, value=estimated_claim_amount)
    initial_ws.cell(row=10, column=5, value=Risk_Adjustment_amount)
    initial_ws.cell(row=10, column=8, value=Outstanding_Claims_amount)

    for i, value in enumerate(discount_rates):
        initial_ws.cell(row=13, column=2 + i, value=float(value))
    for i in range(payment_year):
        initial_ws.cell(row=14, column=2 + i, value=i + 1)
    for i, value in enumerate(payment_patterns):
        initial_ws.cell(row=17, column=2 + i, value=float(value))

    # Add Interest Accretion and CFV values
    Interest_Accretion_values = [entry.get() for entry in Interest_accretion_entries]
    for i, value in enumerate(Interest_Accretion_values):
        initial_ws.cell(row=25, column=2 + i, value=value)
    CFV_values = [entry.get() for entry in CFV_entries]
    for i, value in enumerate(CFV_values):
        initial_ws.cell(row=24, column=2 + i, value=value)

    # Add total values
    initial_ws.cell(row=27, column=2, value=total_cfv)
    initial_ws.cell(row=28, column=2, value=total_interest_accretion)

    # Save the Excel workbook
    wb.save(excel_file_path)
    print("Data exported successfully to Excel.")


def generate_year():
    for spinbox in year_amount_spinbox:
        spinbox.destroy()
    year_amount_spinbox.clear()

    for spinbox in year_risk_spinboxs:
        spinbox.destroy()
    year_risk_spinboxs.clear()

    for spinbox in year_out_spinboxs:
        spinbox.destroy()
    year_out_spinboxs.clear()

    for entry in amount_entrys:
        entry.destroy()
    amount_entrys.clear()

    for entry in amount_risk_entrys:
        entry.destroy()
    amount_risk_entrys.clear()

    for entry in amount_out_entrys:
        entry.destroy()
    amount_out_entrys.clear()

    # Convert year to an integer
    year = int(year_spinbox.get())

    for x in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=2 + x, column=0)
        year_amount_spinbox.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=2 + x, column=1)
        amount_entrys.append(entry)

    for y in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=8 + y, column=0)
        year_risk_spinboxs.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=8 + y, column=1)
        amount_risk_entrys.append(entry)

    for a in range(year):
        spinbox = tk.Spinbox(user_info_frame, from_=2020, to=2050, width=10)
        spinbox.grid(row=14 + a, column=0)
        year_out_spinboxs.append(spinbox)

        entry = tk.Entry(user_info_frame, width=10)
        entry.grid(row=14 + a, column=1)
        amount_out_entrys.append(entry)


def calculate_totals():
    # Calculate total for amount_entrys
    total_amount = 0
    for entry in amount_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_amount += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Calculate total for amount_out_entrys
    total_outstanding = 0
    for entry in amount_out_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_outstanding += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Calculate total for amount_risk_entrys
    total_risk = 0
    for entry in amount_risk_entrys:
        try:
            value = float(entry.get()) if entry.get() else 0
            total_risk += value
        except ValueError:
            pass  # Ignore non-numeric values

    # Update each entry with the calculated totals
    Estimated_C_amount_entry.delete(0, tk.END)
    Estimated_C_amount_entry.insert(0, str(total_amount))

    Risk_Adjustment_entry.delete(0, tk.END)
    Risk_Adjustment_entry.insert(0, str(total_risk))

    Outstanding_Claims_entry.delete(0, tk.END)
    Outstanding_Claims_entry.insert(0, str(total_outstanding))


def generate_payment_pattern_spinboxes():
    """Generate spinboxes for payment patterns based on the selected payment year and dynamically create CFV and Interest Accretion entries."""
    # Clear previous spinboxes and entries
    for spinbox in payment_patterns_spinboxes:
        spinbox.destroy()
    payment_patterns_spinboxes.clear()

    for entry in CFV_entries:
        entry.destroy()
    CFV_entries.clear()

    for entry in Interest_accretion_entries:
        entry.destroy()
    Interest_accretion_entries.clear()

    payment_year = int(payment_year_spinbox.get())

    for x in range(payment_year):
        spinbox = tk.Spinbox(
            user_info_frame,
            from_=0.0,
            to=1.0,
            increment=0.0001,
            width=10,
            format="%.4f",
            command=update_cfv_and_interest_accretion,
        )
        spinbox.grid(row=26, column=1 + x)
        payment_patterns_spinboxes.append(spinbox)

        # Create corresponding CFV entry box
        cfv_entry = tk.Entry(user_info_frame, width=20)
        cfv_entry.grid(row=27, column=1 + x)
        CFV_entries.append(cfv_entry)
        cfv_entry.config(state="readonly")

        # Create corresponding Interest Accretion entry box
        interest_accretion_entry = tk.Entry(user_info_frame, width=20)
        interest_accretion_entry.grid(row=28, column=1 + x)
        Interest_accretion_entries.append(interest_accretion_entry)
        interest_accretion_entry.config(state="readonly")

    # Also generate discount rate spinboxes when payment years are updated
    generate_discount_rate_spinboxes()


def next_interface():
    try:
        # Export the current values to Excel
        total_cfv = total_cfv_entry.get()
        total_interest_accretion = total_interest_accretion_entry.get()

        # Get the Excel file path
        excel_file_path = r"C:\GI_Automation\output\user_data.xlsx"

        # Load the workbook
        wb = openpyxl.load_workbook(excel_file_path)

        # Get the User Data sheet
        if "User Data" in wb.sheetnames:
            user_data_sheet = wb["User Data"]

            # Add SUM formulas to column Q (17)
            user_data_sheet.cell(row=2, column=17, value="=SUM(A2:P2)")  # For Total CFV
            user_data_sheet.cell(row=3, column=17, value="=SUM(A3:P3)")

            if isinstance(user_data_sheet, (float)):
                print()
            else:
                print("The value is not a number or has not been calculated.")  # For Total Interest Accretion

            # Optionally add header and formatting

            # Save the workbook
            wb.save(excel_file_path)
            wb.close()

        # Export current values to Excel
        export_to_excel(total_cfv, total_interest_accretion)

        # Close current interface
        interface.destroy()

        # Launch next interface
        subprocess.Popen(["python", r"C:\GI_Automation\GUI\interface3_1.py"])

    except Exception as e:
        print(f"Error in next_interface: {e}")


def back():
    # Store user_data before going back
    interface.destroy()
    subprocess.Popen(
        ["python", r"C:\GI_Automation\GUI\main.py"]
    )


# Setup main window
interface = tk.Tk()
interface.title("Interface2")
interface.geometry("2500x1000")


frame = tk.Frame(interface)
frame.pack()

user_info_frame = tk.LabelFrame(frame, text="User Information", padx=10, pady=10)
user_info_frame.grid(row=0, column=0, padx=10, pady=10)


product_name_lable = tk.Label(user_info_frame, text="product :")
product_name_lable.grid(row=1, column=0, sticky="w")


product_name_entry = tk.Entry(user_info_frame, width=20)
product_name_entry.grid(row=1, column=1, pady=5)

year_lable = tk.Label(user_info_frame, text="Year:")
year_lable.grid(row=1, column=3, sticky="w")

year_spinbox = tk.Spinbox(user_info_frame, from_=1, to=5, command=generate_year)
year_spinbox.grid(row=1, column=4, pady=5)

# Estimated Claim Amount
Estimated_C_amount_label = tk.Label(user_info_frame, text="Estimated Claim Amount (IBNR):")
Estimated_C_amount_label.grid(row=7, column=0, sticky="w")

Risk_Adjustment_label = tk.Label(user_info_frame, text="Risk Adjustment:")
Risk_Adjustment_label.grid(row=13, column=0, sticky="w")

Outstanding_Claims_label = tk.Label(user_info_frame, text="Outstanding Claims:")
Outstanding_Claims_label.grid(row=20, column=0, sticky="w")


Estimated_C_amount_entry = tk.Entry(user_info_frame, width=20)
Estimated_C_amount_entry.grid(row=7, column=1, pady=5)
Risk_Adjustment_entry = tk.Entry(user_info_frame, width=20)
Risk_Adjustment_entry.grid(row=13, column=1, pady=5)
Outstanding_Claims_entry = tk.Entry(user_info_frame, width=20)
Outstanding_Claims_entry.grid(row=20, column=1, pady=5)


generate_button = tk.Button(user_info_frame, text="Generate calculation ", command=calculate_totals)
generate_button.grid(row=20, column=2, pady=10)

Payment_year_label = tk.Label(user_info_frame, text="Payment Year:")
Payment_year_label.grid(row=23, column=0, sticky="w")

payment_year_spinbox = tk.Spinbox(user_info_frame, from_=1, to=12, command=generate_payment_pattern_spinboxes)
payment_year_spinbox.grid(row=23, column=1, pady=5)

# Discount Rate (dynamic spinboxes will be created)
Discount_rate_label = tk.Label(user_info_frame, text="Discount Rate (%):")
Discount_rate_label.grid(row=24, column=0, sticky="w")

# Estimated Payment Pattern
E_pay_label = tk.Label(user_info_frame, text="Estimated Payment Pattern (%):")
E_pay_label.grid(row=26, column=0, sticky="w")

# Cash Flow Value (CFV) Label
CFV_label = tk.Label(user_info_frame, text="Cash Flow Value (CFV):")
CFV_label.grid(row=27, column=0, sticky="w")  # CFV label

# Interest Accretion
Interest_accretion_label = tk.Label(user_info_frame, text="Interest Accretion:")
Interest_accretion_label.grid(row=28, column=0, sticky="w")

# Calculation Button
calculate_button = tk.Button(user_info_frame, text="Calculate", command=calculate_discount_factors)
calculate_button.grid(row=35, column=0, pady=10)

export_button = tk.Button(
    user_info_frame,
    text="Export to Excel",
    command=lambda: export_to_excel(total_cfv_entry.get(), total_interest_accretion_entry.get()),
)
export_button.grid(row=35, column=3, pady=20)


total_cfv_label = tk.Label(user_info_frame, text="Total CFV:")
total_cfv_label.grid(row=29, column=0, sticky="w")

# Entry box for Total CFV
total_cfv_entry = tk.Entry(user_info_frame, width=20)
total_cfv_entry.grid(row=29, column=1)
total_cfv_entry.config(state="readonly")

total_interest_accretion_label = tk.Label(user_info_frame, text="Total Interest Accretion:")
total_interest_accretion_label.grid(row=30, column=0, sticky="w")

# Entry box for Total Interest Accretion
total_interest_accretion_entry = tk.Entry(user_info_frame, width=20)
total_interest_accretion_entry.grid(row=30, column=1)
total_interest_accretion_entry.config(state="readonly")

# Next and Back Buttons
next_button = tk.Button(user_info_frame, text="Next", command=next_interface)
next_button.grid(row=35, column=1, pady=10)

# Back Button
back_button = tk.Button(user_info_frame, text="Back", command=back)
back_button.grid(row=35, column=2, pady=10)


interface.mainloop()
